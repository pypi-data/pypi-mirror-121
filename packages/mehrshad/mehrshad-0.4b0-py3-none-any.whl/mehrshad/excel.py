########## Wrote by Mehrshad! ##########
import os
import subprocess

from win32com import client

try:
    import pathlib
except ModuleNotFoundError:
    subprocess.call(['pip', 'install', 'pathlib'])
    import pathlib

try:
    from openpyxl import *
    from openpyxl.styles import Border, Font, Side
except ModuleNotFoundError:
    subprocess.call(['pip', 'install', 'openpyxl'])
    from openpyxl import *
    from openpyxl.styles import Border, Font, Side

defaultFileName = ''
defaultFilePath = ''

B_Titr = Font(name='B Titr', charset=178, family=None, b=False, i=False, strike=None, outline=None,
              shadow=None, condense=None, color=None, extend=None, sz=14.0, u=None, vertAlign=None, scheme=None)

B_Nazanin = Font(name='B Nazanin', charset=178, family=None, b=False, i=False, strike=None, outline=None,
                 shadow=None, condense=None, color=None, extend=None, sz=16.0, u=None, vertAlign=None, scheme=None)

B_Nazanin_bold = Font(name='B Nazanin', charset=178, family=None, b=True, i=False, strike=None, outline=None,
                      shadow=None, condense=None, color=None, extend=None, sz=16.0, u=None, vertAlign=None, scheme=None)

B_Titr_bold = Font(name='B Titr', charset=178, family=None, b=True, i=False, strike=None, outline=None,
                   shadow=None, condense=None, color=None, extend=None, sz=14.0, u=None, vertAlign=None, scheme=None)

normalBorder = Border(outline=True, diagonalUp=False, diagonalDown=False, start=None, end=None,
                      left=Side(style='double', color=None), right=Side(style='double', color=None),
                      top=Side(style='dotted', color=None), bottom=Side(style='dotted', color=None)
                      )


def read_sheet(sheetName: str, fileName='', path=''):
    global defaultFileName, defaultFilePath

    if fileName == '':
        fileName = defaultFileName
    if path == '':
        path = defaultFilePath
    if path != '' and path[-1] != '/':
        path += '/'

    address = f'{path}{defaultFileName}'
    if address[-5:] != '.xlsx':
        address += '.xlsx'

    file = load_workbook(address)
    sheet = file[sheetName]
    return sheet


def set_header(sheetName: str, text='Title', font='B Titr', fontSize=14, fileName='', path=''):
    global defaultFileName, defaultFilePath

    if fileName == '':
        fileName = defaultFileName
    if path == '':
        path = defaultFilePath
    if path != '' and path[-1] != '/':
        path += '/'

    address = f'{path}{defaultFileName}'
    if address[-5:] != '.xlsx':
        address += '.xlsx'

    file = load_workbook(address)
    sheet = file[sheetName]
    sheet.oddHeader.center.text = text
    sheet.oddHeader.center.size = fontSize
    sheet.oddHeader.center.font = font
    file.save(address)
    return 'done!'


def set_footer(sheetName: str, text='Title', font='B Titr', fontSize=14, fileName='', path=''):
    global defaultFileName, defaultFilePath

    if fileName == '':
        fileName = defaultFileName
    if path == '':
        path = defaultFilePath
    if path != '' and path[-1] != '/':
        path += '/'

    address = f'{path}{defaultFileName}'
    if address[-5:] != '.xlsx':
        address += '.xlsx'

    file = load_workbook(address)
    sheet = file[sheetName]
    sheet.oddFooter.center.text = text
    sheet.oddFooter.center.size = fontSize
    sheet.oddFooter.center.font = font
    file.save(address)
    return 'done!'


def get_cell_value(sheetName: str, cell: str, fileName='', path=''):
    global defaultFileName, defaultFilePath

    if fileName == '':
        fileName = defaultFileName
    if path == '':
        path = defaultFilePath
    if path != '' and path[-1] != '/':
        path += '/'

    address = f'{path}{defaultFileName}'
    if address[-5:] != '.xlsx':
        address += '.xlsx'

    file = load_workbook(address)
    sheet = file[sheetName]
    return sheet[cell.upper()].value


def set_cell_value(sheetName: str, cell: str, value, setFont=True, font='B Titr', setBorder=True, bold=False, fileName='', path='', customFont=None):
    global defaultFileName, defaultFilePath, B_Titr, B_Nazanin, B_Titr_bold, B_Nazanin_bold, normalBorder

    if fileName == '':
        fileName = defaultFileName
    if path == '':
        path = defaultFilePath
    if path != '' and path[-1] != '/':
        path += '/'

    address = f'{path}{fileName}'
    if address[-5:] != '.xlsx':
        address += '.xlsx'

    file = load_workbook(address)
    sheet = file[sheetName]
    sheet[cell.upper()] = value

    if setFont and customFont == None:
        if bold:
            if font == 'B Nazanin':
                sheet[cell.upper()].font = B_Nazanin_bold
            elif font == 'B Titr':
                sheet[cell.upper()].font = B_Titr_bold
        elif font == 'B Nazanin':
            sheet[cell.upper()].font = B_Nazanin
        elif font == 'B Titr':
            sheet[cell.upper()].font = B_Titr
    elif customFont != None:
        sheet[cell.upper()].font = customFont

    if setBorder:
        sheet[cell.upper()].border = normalBorder
    file.save(address)
    return 'done!'


def merge_cells(sheetName: str, cells='A1:B2', text='', fileName='', path=''):
    global defaultFileName, defaultFilePath

    if fileName == '':
        fileName = defaultFileName
    if path == '':
        path = defaultFilePath
    if path != '' and path[-1] != '/':
        path += '/'

    address = f'{path}{defaultFileName}'
    if address[-5:] != '.xlsx':
        address += '.xlsx'

    file = load_workbook(address)
    sheet = file[sheetName]
    sheet.merge_cells(cells.upper())
    cell = cells[:cells.find(':')].upper()
    sheet[cell] = text
    sheet[cell].font = B_Titr
    sheet[cell].border = normalBorder
    file.save(address)
    return 'done!'


def convert_to_pdf(outputName='output', pages={1: 'A1:G20'}, saveLocation='desktop', fileName='', path=str(pathlib.Path().absolute())):
    global defaultFileName

    if fileName == '':
        fileName = defaultFileName
    if path == str(pathlib.Path().absolute()):
        path += '/'+defaultFilePath
    if path[-1] != '/':
        path += '/'

    address = f'{path}{defaultFileName}'
    if address[-5:] != '.xlsx':
        address += '.xlsx'

    excel = client.DispatchEx("Excel.Application")
    excel.Visible = False

    if saveLocation.lower() == 'downloads':
        savePath = os.path.join(os.path.join(
            os.environ['USERPROFILE']), 'Downloads')
    elif saveLocation.lower() == 'documents':
        savePath = os.path.join(os.path.join(
            os.environ['USERPROFILE']), 'Documents')
    else:
        savePath = os.path.join(os.path.join(
            os.environ['USERPROFILE']), 'Desktop')

    savePath = f'{savePath}\\{outputName}.pdf'

    wb = excel.Workbooks.Open(address)
    for index in pages.keys():
        ws = wb.Worksheets[index - 1]
        ws.PageSetup.Zoom = False
        ws.PageSetup.FitToPagesTall = 1
        ws.PageSetup.FitToPagesWide = 1
        if pages[index] != None:
            ws.PageSetup.PrintArea = str(pages[index]).upper()

    index = list(pages.keys())
    wb.WorkSheets(index).Select()
    wb.ActiveSheet.ExportAsFixedFormat(0, savePath)
    wb.Close(SaveChanges=0)
    excel.Quit()
    return f'محل ذخیره فایل پی دی اف:\n{savePath}'
