import collections
import logging
import os
import re

from PyQt5 import QtCore, QtGui

import openpyxl

import tabula

import numpy as np

import pandas as pd


class InvalidViewError(Exception):
    """Exception raised for dynamic matrix view related errors.
    """


class DynamicMatrixModel(QtCore.QAbstractTableModel):

    zones = [('A', 'B', 'C', 'D', 'E'), ('A', 'B', 'C', 'D'), ('A', 'B'), ('C', 'D'), ('E',), ('P',), ('Z',)]

    def __init__(self, *args, **kwargs):
        """Constructor.
        """

        super(DynamicMatrixModel, self).__init__(*args, **kwargs)

        self._dynamic_matrix = pd.DataFrame()

    def clear(self):
        """Clear the dynamic matrix.
        """

        self._dynamic_matrix = pd.DataFrame()

        self.layoutChanged.emit()

    def columnCount(self, parent=None):
        """Return the number of columns of the model for a given parent.

        Returns:
            int: the number of columns
        """

        return self._dynamic_matrix.shape[1]

    def data(self, index, role):
        """Get the data at a given index for a given role.

        Args:
            index (QtCore.QModelIndex): the index
            role (int): the role

        Returns:
            QtCore.QVariant: the data
        """

        if not index.isValid():
            return QtCore.QVariant()

        if self._dynamic_matrix.empty:
            return QtCore.QVariant()

        row = index.row()
        col = index.column()

        if role == QtCore.Qt.DisplayRole:
            return str(self._dynamic_matrix.iloc[row, col])
        elif role == QtCore.Qt.ToolTipRole:
            return str(self._dynamic_matrix.iloc[row, col])

    @property
    def dynamic_matrix(self):
        """Return the dynamic matrix.

        Returns:
            pd.DataFrame: the dynamic matrix
        """

        return self._dynamic_matrix

    @dynamic_matrix.setter
    def dynamic_matrix(self, matrix):
        """Set the dynamic matrix.

        Args:
            matrix (pd.DataFrame): the dynamic matrix
        """

        self._dynamic_matrix = matrix

        self.layoutChanged.emit()

    def export(self, workbook, gene):
        """Export the raw data to an excel spreadsheet.

        Args:
            workbook (openpyxl.workbook.workbook.Workbook): the workbook
        """

        worksheet = workbook.create_sheet(gene)

        # Create a worksheet for the dynamic matrix
        comp = 1
        worksheet.cell(row=comp, column=1).value = 'Dynamic matrix'

        comp += 1

        for i, v in enumerate(self._dynamic_matrix.columns):
            worksheet.cell(row=comp, column=i+2).value = v

        for i, v in enumerate(self._dynamic_matrix.index):
            comp += 1
            worksheet.cell(row=comp, column=1).value = v
            for j in range(len(self._dynamic_matrix.columns)):
                worksheet.cell(row=comp, column=j+2).value = str(self._dynamic_matrix.iloc[i, j])

        all_zones = [''.join(z) for z in DynamicMatrixModel.zones]

        comp += 2
        worksheet.cell(row=comp, column=1).value = 'Averages'
        averages = self.get_averages(all_zones)

        comp += 1
        for i, v in enumerate(averages.columns):
            worksheet.cell(row=comp, column=i+2).value = v

        for i, v in enumerate(averages.index):
            comp += 1
            worksheet.cell(row=comp, column=1).value = v
            for j in range(len(averages.columns)):
                worksheet.cell(row=comp, column=j+2).value = averages.iloc[i, j]

        comp += 2
        worksheet.cell(row=comp, column=1).value = 'Std Devs'
        stds = self.get_stds(all_zones)

        comp += 1
        for i, v in enumerate(stds.columns):
            worksheet.cell(row=comp, column=i+2).value = v

        for i, v in enumerate(stds.index):
            comp += 1
            worksheet.cell(row=comp, column=1).value = v
            for j in range(len(stds.columns)):
                worksheet.cell(row=comp, column=j+2).value = stds.iloc[i, j]

        comp += 2
        worksheet.cell(row=comp, column=1).value = 'N values'
        n_values = self.get_n_values(all_zones)

        comp += 1
        for i, v in enumerate(n_values.columns):
            worksheet.cell(row=comp, column=i+2).value = v

        for i, v in enumerate(n_values.index):
            comp += 1
            worksheet.cell(row=comp, column=1).value = v
            for j in range(len(n_values.columns)):
                worksheet.cell(row=comp, column=j+2).value = n_values.iloc[i, j]

    def get_averages(self, zones):
        """Getter for the averages of each entry of the dynamic matrix for each selected zone.

        Args:
        Returns:
            pandas.DataFrame: the averages
        """

        filtered_zones = []
        for zone in zones:
            if zone not in self._dynamic_matrix.index:
                continue
            filtered_zones.append(zone)

        if not filtered_zones:
            return pd.DataFrame()

        samples = self._dynamic_matrix.columns

        averages = pd.DataFrame(np.nan, index=filtered_zones, columns=samples)

        for z in filtered_zones:
            for s in samples:
                averages.loc[z, s] = np.mean(self._dynamic_matrix.loc[z, s])

        averages = averages.round(3)

        return averages

    def get_diff(self, zones):
        """Getter for the difference between the max and the min of each dynamic matrix entry for each selected zone.

        Args:
        Returns:
            pandas.DataFrame: the difference matrix
        """

        filtered_zones = []
        for zone in zones:
            if zone not in self._dynamic_matrix.index:
                continue
            filtered_zones.append(zone)

        if not filtered_zones:
            return pd.DataFrame()

        samples = self._dynamic_matrix.columns

        diff = pd.DataFrame(np.nan, index=filtered_zones, columns=samples)

        for z in filtered_zones:
            for s in samples:
                if self._dynamic_matrix.loc[z, s]:
                    diff.loc[z, s] = max(self._dynamic_matrix.loc[z, s]) - min(self._dynamic_matrix.loc[z, s])
                else:
                    diff.loc[z, s] = np.nan

        diff = diff.round(3)

        return diff

    def get_n_values(self, zones):
        """Getter for the matrix which stores the number of samples of each entry of the dynamic matrix for each selected zone.

        Args:
            zones (list of str): the zones

        Returns:
            pandas.DataFrame: the matrix.
        """

        filtered_zones = []
        for zone in zones:
            if zone not in self._dynamic_matrix.index:
                continue
            filtered_zones.append(zone)

        samples = self._dynamic_matrix.columns

        n_values = pd.DataFrame(np.nan, index=filtered_zones, columns=samples)

        for z in filtered_zones:
            for s in samples:
                n_values.loc[z, s] = len(self._dynamic_matrix.loc[z, s])

        return n_values

    def get_stds(self, zones):
        """Getter for the standard deviation matrix of each entry of the dynamic matrix for each selected zone.

        Args:
            zones (list of str): the zones

        Returns:
            pandas.DataFrame: the matrix of standard deviations
        """

        # Keep only the zones which are present as an index of the current dynamic matrix
        filtered_zones = []
        for zone in zones:
            if zone not in self._dynamic_matrix.index:
                continue
            filtered_zones.append(zone)

        # Fetch the sample names
        samples = self._dynamic_matrix.columns

        # Initialize the standard deviation data frame
        stds = pd.DataFrame(np.nan, index=filtered_zones, columns=samples)

        # Compute the standard deviation
        for z in filtered_zones:
            for s in samples:
                stds.loc[z, s] = np.std(self._dynamic_matrix.loc[z, s])

        stds = stds.round(3)

        return stds

    def headerData(self, idx, orientation, role):
        """Returns the header data for a given index, orientation and role.

        Args:
            idx (int): the index
            orientation (int): the orientation
            role (int): the role
        """

        if role == QtCore.Qt.DisplayRole:
            if orientation == QtCore.Qt.Horizontal:
                return self._dynamic_matrix.columns[idx]
            else:
                return self._dynamic_matrix.index[idx]

    def rowCount(self, parent=None):
        """Return the number of rows of the model.

        Returns:
            int: the number of rows
        """

        return self._dynamic_matrix.shape[0]
