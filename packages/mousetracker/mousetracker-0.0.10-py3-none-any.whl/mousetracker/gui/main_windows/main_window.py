"""This modules implements the following classes:
    - MainWindow
"""

import copy
import logging
import os
import sys

import yaml

from openpyxl import load_workbook

from PyQt5 import QtCore, QtGui, QtWidgets

import mousetracker
from mousetracker.__pkginfo__ import __version__
from mousetracker.gui.widgets.groups_widget import GroupsWidget
from mousetracker.gui.widgets.logger_widget import QTextEditLogger
from mousetracker.gui.widgets.statistics_widget import StatisticsWidget
from mousetracker.kernel.models.excel_files_model import ExcelFilesModel, ExcelFileModelError
from mousetracker.kernel.models.groups_model import GroupsModel
from mousetracker.kernel.models.mouse_monitoring_model import MouseMonitoringModel
from mousetracker.kernel.utils.progress_bar import progress_bar


class MainWindow(QtWidgets.QMainWindow):
    """This class implements the main window of the application.
    """

    set_groups_model = QtCore.pyqtSignal(GroupsModel, list)

    set_properties = QtCore.pyqtSignal(list)

    def __init__(self, parent=None):
        """Constructor.

        Args:
            parent (QtCore.QObject): the parent window
        """

        super(MainWindow, self).__init__(parent)

        self._init_ui()

    def _build_events(self):
        """Build the signal/slots.
        """

        self._excel_files_listview.selectionModel().selectionChanged.connect(self.on_select_excel_file)
        self.set_groups_model.connect(self._groups_widgets.on_set_groups_model)
        self.set_properties.connect(self._groups_widgets.on_set_properties)
        self._groups_widgets.compute_statistics.connect(self.on_build_statistics_widget)
        self._tabs.tabCloseRequested.connect(lambda index: self._tabs.removeTab(index))

    def _build_layout(self):
        """Build the layout.
        """

        main_layout = QtWidgets.QVBoxLayout()

        hlayout = QtWidgets.QHBoxLayout()

        hlayout.addWidget(self._excel_files_listview)

        hlayout.addWidget(self._tabs, stretch=2)

        main_layout.addLayout(hlayout, stretch=2)

        main_layout.addWidget(self._logger.widget, stretch=1)

        self._main_frame.setLayout(main_layout)

    def _build_menu(self):
        """Build the menu.
        """

        menubar = self.menuBar()

        file_menu = menubar.addMenu('&File')

        file_action = QtWidgets.QAction('&Open mousetracker files', self)
        file_action.setShortcut('Ctrl+O')
        file_action.setStatusTip('Open mousetracker files')
        file_action.triggered.connect(self.on_open_mousetracker_files)
        file_menu.addAction(file_action)

        file_menu.addSeparator()

        exit_action = QtWidgets.QAction('&Exit', self)
        exit_action.setShortcut('Ctrl+Q')
        exit_action.setStatusTip('Exit mousetracker')
        exit_action.triggered.connect(self.on_quit_application)
        file_menu.addAction(exit_action)

        group_menu = menubar.addMenu('&Groups')

        export_group_action = QtWidgets.QAction('&Export groups', self)
        export_group_action.setShortcut('Ctrl+E')
        export_group_action.setStatusTip('Export groups')
        export_group_action.triggered.connect(self.on_export_groups)
        group_menu.addAction(export_group_action)

        group_menu.addSeparator()

        import_group_action = QtWidgets.QAction('&Import groups', self)
        import_group_action.setShortcut('Ctrl+I')
        import_group_action.setStatusTip('Import groups')
        import_group_action.triggered.connect(self.on_import_groups)
        group_menu.addAction(import_group_action)

    def _build_widgets(self):
        """Build the widgets.
        """

        self._main_frame = QtWidgets.QFrame(self)

        self._excel_files_listview = QtWidgets.QListView(self)
        self._excel_files_listview.setSelectionMode(QtWidgets.QListView.SingleSelection)
        self._excel_files_listview.setModel(ExcelFilesModel(self))

        self._tabs = QtWidgets.QTabWidget(self)
        self._tabs.setTabsClosable(True)

        self._excel_file_contents_tableview = QtWidgets.QTableView(self)
        self._excel_file_contents_tableview.setModel(MouseMonitoringModel(self))

        self._groups_widgets = GroupsWidget(self)

        self._tabs.addTab(self._excel_file_contents_tableview, 'Data')
        self._tabs.addTab(self._groups_widgets, 'Groups')

        self._tabs.tabBar().setTabButton(0, QtWidgets.QTabBar.RightSide, None)
        self._tabs.tabBar().setTabButton(1, QtWidgets.QTabBar.RightSide, None)

        self._logger = QTextEditLogger(self)
        self._logger.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s'))
        logging.getLogger().addHandler(self._logger)
        logging.getLogger().setLevel(logging.INFO)

        self.setCentralWidget(self._main_frame)

        self.setGeometry(0, 0, 1200, 1100)

        self.setWindowTitle('mousetracker {}'.format(__version__))

        self._progress_label = QtWidgets.QLabel('Progress')
        self._progress_bar = QtWidgets.QProgressBar()
        progress_bar.set_progress_widget(self._progress_bar)
        self.statusBar().showMessage('mousetracker {}'.format(__version__))
        self.statusBar().addPermanentWidget(self._progress_label)
        self.statusBar().addPermanentWidget(self._progress_bar)

        icon_path = os.path.join(mousetracker.__path__[0], 'icons', 'mousetracker.png')
        self.setWindowIcon(QtGui.QIcon(icon_path))

        self.show()

    def _init_ui(self):
        """Initializes the ui.
        """

        self._build_widgets()

        self._build_layout()

        self._build_menu()

        self._build_events()

    @property
    def excel_files_listview(self):

        return self._excel_files_listview

    def on_build_statistics_widget(self, selected_property, groups_model):
        """
        """

        if groups_model.rowCount() == 0:
            logging.warning('No groups defined')
            return

        statistics_widget = StatisticsWidget(selected_property, groups_model, self)
        self._tabs.addTab(statistics_widget, 'Statistics')
        self._tabs.setTabsClosable(True)

    def on_export_groups(self):
        """Export groups.
        """

        yaml_file = QtWidgets.QFileDialog.getSaveFileName(self, caption='Export group as ...', filter="YAML files (*.yaml *.yml)")
        if not yaml_file:
            return

        yaml_file = yaml_file[0]
        if not yaml_file:
            return

        excel_files_model = self._excel_files_listview.model()

        exportable_data = []
        for r in range(excel_files_model.rowCount()):
            e_index = excel_files_model.index(r, 0)
            excel_file = excel_files_model.data(e_index, QtCore.Qt.DisplayRole)
            groups_model = excel_files_model.data(e_index, ExcelFilesModel.group_model)

            groups = sorted(groups_model.groups, key=lambda x: x[0])

            exported_groups = []
            for group, model, selected in groups:
                contents = [model.data(model.index(i, 0), QtCore.Qt.DisplayRole) for i in range(model.rowCount())]
                exported_groups.append((group, contents, selected))

            exportable_data.append({'excel_file': excel_file, 'groups': exported_groups})

        with open(yaml_file, 'w') as f:
            yaml.dump(exportable_data, f)

    def on_import_groups(self):
        """Import groups.
        """

        # Pop up a file browser for selecting the workbooks
        yaml_file = QtWidgets.QFileDialog.getOpenFileName(self, 'Import groups', '', 'YAML Files (*.yml *.yaml)')
        if not yaml_file:
            return

        yaml_file = yaml_file[0]
        if not yaml_file:
            return

        with open(yaml_file, 'r') as f:
            imported_groups = yaml.load(f, Loader=yaml.FullLoader)

        excel_files_model = self._excel_files_listview.model()
        for group_dict in imported_groups:

            excel_file = group_dict['excel_file']

            try:
                excel_files_model.add_excel_file(excel_file)
            except ExcelFileModelError as e:
                logging.error(str(e))
                continue

            row = excel_files_model.rowCount() - 1
            e_index = excel_files_model.index(row, 0)
            groups_model = excel_files_model.data(e_index, ExcelFilesModel.group_model)
            for group_name, group_contents, selected in group_dict['groups']:
                groups_model.add_group(group_name, selected=selected)
                g_row = groups_model.rowCount() - 1
                g_index = groups_model.index(g_row, 0)
                group_contents_model = groups_model.data(g_index, GroupsModel.model)
                for item in group_contents:
                    group_contents_model.add_item(item)

    def on_open_mousetracker_files(self):
        """Event handler which opens a dialog for selecting data files.
        """

        # Pop up a file browser for selecting the workbooks
        excel_files = QtWidgets.QFileDialog.getOpenFileNames(self, 'Open data files', '', 'Data Files (*.xls *.xlsx)')[0]
        if not excel_files:
            return

        n_excel_files = len(excel_files)
        progress_bar.reset(n_excel_files)

        n_loaded_files = 0

        excel_files_model = self._excel_files_listview.model()

        # Loop over the pig directories
        for progress, data_file in enumerate(excel_files):

            # Read the pdf file and add the data to the model. Any kind of error must be caught here.
            try:
                self.statusBar().showMessage('Reading {} file ...'.format(data_file))
                excel_files_model.add_excel_file(data_file)

            except Exception as error:
                logging.error(str(error))

            else:
                n_loaded_files += 1

            progress_bar.update(progress+1)

        self.statusBar().showMessage('')
        logging.info('Loaded successfully {} file(s) out of {}'.format(n_loaded_files, n_excel_files))

    def on_quit_application(self):
        """Event handler which quits the application.
        """

        choice = QtWidgets.QMessageBox.question(self, 'Quit', "Do you really want to quit ?", QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No)
        if choice == QtWidgets.QMessageBox.Yes:
            sys.exit()

    def on_select_excel_file(self, selection):
        """Event handler which displays the contents of an excel file.
        """

        indexes = selection.indexes()

        if not indexes:
            return

        excel_files_model = self._excel_files_listview.model()

        data_frame = excel_files_model.data(indexes[0], ExcelFilesModel.data_frame)
        groups_model = excel_files_model.data(indexes[0], ExcelFilesModel.group_model)

        excel_file_contents_model = self._excel_file_contents_tableview.model()
        excel_file_contents_model.set_data_frame(data_frame)

        animal = data_frame.animal
        n_zones = data_frame.n_zones

        animals = data_frame[animal][0::n_zones].to_list()

        self.set_groups_model.emit(groups_model, animals)

        properties = data_frame.properties
        self.set_properties.emit(properties)

    def export(self, filename, selected_properties):
        """
        """

        _, ext = os.path.splitext(filename)

        if ext != '.xlsx':
            logging.error('Invalid file extension. Must be .xlsx')
            return

        # Export the current data to 'data' sheet
        index = self._excel_files_listview.currentIndex()
        excel_files_model = self._excel_files_listview.model()
        dataframe = excel_files_model.data(index, ExcelFilesModel.data_frame)
        dataframe.to_excel(filename, sheet_name='data', index=False)

        workbook = load_workbook(filename=filename)

        groups_model = self._groups_widgets.groups_listview.model()

        # Export the groups
        groups_sheet = workbook.create_sheet('groups')
        comp = 0
        for group, model, selected in groups_model.groups:
            if not selected:
                continue
            comp += 1
            groups_sheet.cell(row=1, column=comp).value = group
            for i in range(model.rowCount()):
                index = model.index(i, 0)
                mouse = model.data(index, QtCore.Qt.DisplayRole)
                groups_sheet.cell(row=i+2, column=comp).value = mouse

        # Export the statistics
        for prop in selected_properties:
            statistics_sheet = workbook.create_sheet('statistics {}'.format(prop))

            all_zones = groups_model.get_statistics_zones()
            statistics = groups_model.get_statistics(prop, all_zones)

            comp = 1
            statistics_sheet.cell(row=comp, column=1).value = 'selected property'
            statistics_sheet.cell(row=comp, column=2).value = prop
            for s in ['mean', 'std', 'n']:
                comp += 1
                statistics_sheet.cell(row=comp, column=1).value = s
                for group, df in statistics[s].items():
                    comp += 1
                    statistics_sheet.cell(row=comp, column=1).value = group
                    # Write the column of the dataframe
                    comp += 1
                    for i, c in enumerate(df.columns):
                        statistics_sheet.cell(row=comp, column=i+2).value = c
                    # Write the index and data of the dataframe
                    for i in range(len(df.index)):
                        comp += 1
                        statistics_sheet.cell(row=comp, column=1).value = df.index[i]
                        for j in range(len(df.columns)):
                            statistics_sheet.cell(row=comp, column=j+2).value = df.iloc[i, j]
                    comp += 1

            # Export the student test to 'student test' sheet
            student_test_sheet = workbook.create_sheet('student tests {}'.format(prop))
            comp = 1
            student_test_sheet.cell(row=comp, column=1).value = 'selected property'
            student_test_sheet.cell(row=comp, column=2).value = prop

            student_tests_zones = groups_model.get_student_tests_zones()
            student_tests = groups_model.get_student_tests(prop, student_tests_zones)

            for zone, df_dict in student_tests.items():
                comp += 1
                student_test_sheet.cell(row=comp, column=1).value = zone
                for day, df in df_dict.items():
                    comp += 1
                    student_test_sheet.cell(row=comp, column=1).value = day
                    # Write the column of the dataframe
                    comp += 1
                    for i, c in enumerate(df.columns):
                        student_test_sheet.cell(row=comp, column=i+2).value = c
                    # Write the index and data of the dataframe
                    for i in range(len(df.index)):
                        comp += 1
                        student_test_sheet.cell(row=comp, column=1).value = df.index[i]
                        for j in range(len(df.columns)):
                            student_test_sheet.cell(row=comp, column=j+2).value = df.iloc[i, j]
                    comp += 1

        try:
            workbook.save(filename)
        except PermissionError as error:
            logging.error(str(error))
            return
