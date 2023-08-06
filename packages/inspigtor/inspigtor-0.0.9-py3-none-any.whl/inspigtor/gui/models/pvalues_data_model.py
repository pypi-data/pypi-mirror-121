"""This module implements:
    - PValuesDataModel
"""

import logging

import openpyxl

from PyQt5 import QtCore, QtGui


class PValuesDataModel(QtCore.QAbstractTableModel):
    """This model stores the p values table coming from Dunn statistical test which produces a matrix of p values.
    """

    def __init__(self, data):
        """Constructor

        Arguments:
            data (pandas.DataFrame): the input p value matrix
        """
        super(PValuesDataModel, self).__init__()
        self._data = data

    def matrix(self):

        return self._data

    def rowCount(self, parent=None):
        """Returns the number of rows of the underlying p value matrix.
        """

        return self._data.shape[0]

    def columnCount(self, parent=None):
        """Returns the number of columns of the underlying p value matrix.
        """

        return self._data.shape[1]

    def data(self, index, role=QtCore.Qt.DisplayRole):
        """Returns the data for the given role and section.

        Arguments:
            idx (int): the section
            role (any Qt role): the role

        Returns:
            any: the data
        """

        if index.isValid():
            if role == QtCore.Qt.DisplayRole:
                return str(self._data.iloc[index.row(), index.column()])
            elif role == QtCore.Qt.ForegroundRole:
                row = index.row()
                column = index.column()
                p_value = self._data.iloc[row, column]
                if p_value < 0.05 and p_value > 0:
                    return QtGui.QBrush(QtCore.Qt.red)

        return None

    def headerData(self, idx, orientation, role):
        """Returns the data for the given role and section in the header with the specified orientation.

        Arguments:
            idx (int): the section
            orientation (QtCore.Qt.Horizontal or QtCore.Qt.Vertical): the orientation
            role (any Qt role): the role

        Returns:
            str: the stringified header data
        """

        if role == QtCore.Qt.DisplayRole:
            return str(self._data.columns[idx]) if orientation == QtCore.Qt.Horizontal else str(self._data.index[idx])

        return None

    def export(self, filename):
        """Export the current model to the given excel file.

        Arguments:
            filename (str): the excel filename
        """

        workbook = openpyxl.Workbook()
        # Remove the first empty sheet created by default
        workbook.remove_sheet(workbook.get_sheet_by_name('Sheet'))

        workbook.create_sheet('p-values')

        worksheet = workbook.get_sheet_by_name('p-values')

        # Write column titles
        for col in range(self.columnCount()):
            col_name = self.headerData(col, QtCore.Qt.Horizontal, role=QtCore.Qt.DisplayRole)
            worksheet.cell(row=1, column=col+2).value = col_name

        # Write row titles
        for row in range(self.rowCount()):
            row_name = self.headerData(row, QtCore.Qt.Vertical, role=QtCore.Qt.DisplayRole)
            worksheet.cell(row=row+2, column=1).value = row_name

        for row in range(self.rowCount()):
            for col in range(self.columnCount()):
                index = self.index(row, col)
                data = self.data(index, QtCore.Qt.DisplayRole)
                worksheet.cell(row=row+2, column=col+2).value = data

        try:
            workbook.save(filename)
        except PermissionError as error:
            logging.error(str(error))
            return
