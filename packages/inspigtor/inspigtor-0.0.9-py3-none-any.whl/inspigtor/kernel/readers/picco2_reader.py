import collections
from datetime import datetime
import logging
import os
import sys

import openpyxl

import numpy as np

import pandas as pd

from inspigtor.kernel.utils.stats import statistical_functions


class PiCCO2FileReaderError(Exception):
    """Exception for PiCCO2 file reader.
    """


class PiCCO2FileReader:
    """This class implements the PiCCO2 device reader.

    This is the base class of inspigtor application. It reads and parses a PiCCO2 file and computes statistics on
    the properties stored in the file (columns). To be read properly, the file must contain a cell with the starting
    time and ending time of the experiment. The t_initial time will be used to define a t_initial - 10 minutes time starting from
    which records intervals will be computed. Those record intervals are those interval on which the average and std
    of a given property are computed. The t_final time will be used to compute pre-mortem statistics.
    """

    def __init__(self, filename):
        """Constructor

        Args:
            filename (str): the PiCCO2 input file
        """

        if not os.path.exists(filename):
            raise PiCCO2FileReaderError('The picco file {} does not exist'.format(filename))

        self._filename = filename

        csv_file = open(self._filename, 'r')

        # Skip the first line, just comments about the device
        csv_file.readline()

        # Read the second line which contains the titles of the general parameters
        line = csv_file.readline().strip()
        line = line[:-1] if line.endswith(';') else line
        general_info_fields = [v.strip() for v in line.split(';')]

        # Read the third line which contains the values of the general parameters
        line = csv_file.readline().strip()
        line = line[:-1] if line.endswith(';') else line
        general_info = [v.strip() for v in line.split(';')]

        # Create a dict out of those parameters
        general_info_dict = collections.OrderedDict(zip(general_info_fields, general_info))

        if 't_initial' not in general_info_dict:
            raise PiCCO2FileReaderError('Missing t_initial value in the general parameters section.')

        if 't_final' not in general_info_dict:
            raise PiCCO2FileReaderError('Missing t_final value in the general parameters section.')

        # Read the fourth line which contains the titles of the pig id parameters
        line = csv_file.readline().strip()
        line = line[:-1] if line.endswith(';') else line
        pig_id_fields = [v.strip() for v in line.split(';') if v.strip()]

        # Read the fifth line which contains the values of the pig id parameters
        line = csv_file.readline().strip()
        line = line[:-1] if line.endswith(';') else line
        pig_id = [v.strip() for v in line.split(';') if v.strip()]

        # Create a dict outof those parameters
        pig_id_dict = collections.OrderedDict(zip(pig_id_fields, pig_id))

        # Concatenate the pig id parameters dict and the general parameters dict
        self._parameters = {**pig_id_dict, **general_info_dict}

        csv_file.close()

        # Reopen the file to guess size of the header
        csv_file = open(self._filename, 'r')

        header_size = 0
        while not csv_file.readline().startswith('Date;Time'):
            header_size += 1

        csv_file.close()

        # Read the rest of the file as a csv file
        self._data = pd.read_csv(self._filename, sep=';', skiprows=header_size, skipfooter=1, engine='python')

        # For some files, times are not written in chronological order, so sort them before doing anything
        self._data = self._data.sort_values(by=['Time'])

        self._time_fmt = '%H:%M:%S'
        self._exp_start = datetime.strptime(self._data.iloc[0]['Time'], self._time_fmt)

        delta_t = datetime.strptime(general_info_dict['t_initial'], self._time_fmt) - datetime.strptime(self._data['Time'][0], self._time_fmt)
        if delta_t.days < 0 or delta_t.seconds < 600:
            raise PiCCO2FileReaderError('t_initial - 10 minutes is earlier than the beginning of the experiment for file {}.'.format(self._filename))

        # The evaluation of intervals starts at t_zero - 10 minutes (as asked by experimentalists)
        t_minus_10_strptime = datetime.strptime(general_info_dict['t_initial'], self._time_fmt) - datetime.strptime('00:10:00', self._time_fmt)
        t_minus_10_strptime = datetime.strptime(str(t_minus_10_strptime), self._time_fmt)

        self._t_minus_10_index = 0

        valid_t_minus_10 = False
        delta_ts = []
        first = True
        for i, time in enumerate(self._data['Time']):
            delta_t = datetime.strptime(time, self._time_fmt) - t_minus_10_strptime
            # If the difference between the current time and t_zero - 10 is positive for the first time, then record the corresponding
            # index as being the reference time
            if delta_t.days >= 0:
                delta_ts.append(str(delta_t))
                if first:
                    self._t_minus_10_index = i
                    valid_t_minus_10 = True
                    first = False
            else:
                delta_ts.append('-'+str(-delta_t))

        if not valid_t_minus_10:
            raise PiCCO2FileReaderError('Invalid value for t_initial parameters')

        # Add a column to the original data which show the delta t regarding t_zero - 10 minutes
        self._data.insert(loc=2, column='delta_t', value=delta_ts)

        self._record_intervals = []

        self._record = None

    @ property
    def data(self):
        """Property for the data stored in the csv file

        Returns:
            pandas.DataFrame: the data stored in the csv file.
        """

        return self._data

    @ property
    def filename(self):
        """Property for the reader's filename.

        Returns:
            str: the reader's filename.
        """

        return self._filename

    def get_coverages(self, selected_property='APs'):
        """Compute the coverages for a given property.

        The coverage of a property is the ratio between the number of valid values over the total number of values for a given property over a given record interval.

        Args:
            selected_property (str): the selected properrty for which the coverages will be calculated.

        Returns:
            list of float: the coverages for each record interval
        """

        if not self._record_intervals:
            logging.warning('No record intervals defined yet')
            return []

        coverages = []
        # Compute for each record interval the average and standard deviation of the selected property
        for interval in self._record_intervals:
            first_index, last_index = interval
            coverage = 0.0
            for j in range(first_index, last_index):
                # If the value can be casted to a float, the value is considered to be valid
                try:
                    _ = float(self._data[selected_property].iloc[j])
                except ValueError:
                    continue
                else:
                    coverage += 1.0
            coverages.append(100.0*coverage/(last_index-first_index))

        return coverages

    def get_descriptive_statistics(self, selected_property='APs', selected_statistics=None, interval_indexes=None):
        """Compute the statistics for a given property for the current record intervals.

        For each record interval, computes the average and the standard deviation of the selected property.

        Args:
            selected_property (str): the selected property
            selected_statistics (list): the list of statistics to compute
            interval_indexes (list of int): the indexes of the record intervals to select. If None, all the record intervals will be used.

        Returns:
            dict: a dictionary whose keys are the different statistics computed (e;g; average, median ...) and the values are the list of 
            the value of the statistics over record intervals
        """

        if selected_statistics is None:
            selected_statistics = list(statistical_functions.keys())
        else:
            all_statistics = set(statistical_functions.keys())
            selected_statistics = list(all_statistics.intersection(selected_statistics))

        if not selected_statistics:
            raise PiCCO2FileReaderError('Invalid input statistics')

        if selected_property not in list(self._data.columns):
            raise PiCCO2FileReaderError('Property {} is unknown'.format(selected_property))

        # Some record intervals must have been set before
        if not self._record_intervals:
            raise PiCCO2FileReaderError('No record intervals defined yet')

        if interval_indexes is None:
            interval_indexes = list(range(len(self._record_intervals)))

        statistics = {}

        for index in interval_indexes:
            interval = self._record_intervals[index]
            first_index, last_index = interval
            data = []
            for j in range(first_index, last_index):
                try:
                    data.append(float(self._data[selected_property].iloc[j]))
                except ValueError:
                    continue

            statistics.setdefault('intervals', []).append(index)

            if not data:
                statistics.setdefault('data', []).append(None)
                for stat in selected_statistics:
                    statistics.setdefault(stat, []).append(np.nan)
            else:
                statistics.setdefault('data', []).append(data)
                for stat in selected_statistics:
                    statistics.setdefault(stat, []).append(statistical_functions[stat](data))

        return statistics

    def get_t_final_index(self):
        """Return the first index whose time is superior to t_final.
        """

        for index, time in enumerate(self._data['Time']):
            delta_t = datetime.strptime(self._parameters['t_final'], self._time_fmt) - datetime.strptime(time, self._time_fmt)
            if delta_t.days < 0:
                return index

        return len(self._data['Time'])

    def set_record_interval(self, interval):
        """Set the record intervals.

        Args:
            3-tuples: the record interval. 4-tuple of the form (start,end,record).
        """

        t_max = self.get_t_final_index()

        t_minus_10 = datetime.strptime(self._data['Time'].iloc[self._t_minus_10_index], self._time_fmt)

        self._record_intervals = []

        start, end, record = interval

        self._record = int(record)

        # The record is converted from minutes to seconds
        self._record *= 60
        # Convert strptime to timedelta for further use
        start = (datetime.strptime(start, self._time_fmt) - datetime.strptime('00:00:00', self._time_fmt)).seconds
        end = (datetime.strptime(end, self._time_fmt) - datetime.strptime('00:00:00', self._time_fmt)).seconds

        enter_interval = True
        exit_interval = True
        last_record_index = None
        # Loop over the times [t0-10,end] for defining the first and last indexes (included) that falls in the running interval
        for t_index in range(self._t_minus_10_index, t_max):
            delta_t = (datetime.strptime(self._data['Time'].iloc[t_index], self._time_fmt) - t_minus_10).seconds
            # We have not entered yet in the interval, skip.
            if delta_t < start:
                continue
            # We entered in the interval.
            else:
                # We are in the interval
                if delta_t < end:
                    # First time we entered in the interval, record the corresponding index
                    if enter_interval:
                        first_record_index = t_index
                        enter_interval = False
                # We left the interval
                else:
                    # First time we left the interval, record the corresponding index
                    if exit_interval:
                        last_record_index = t_index
                        exit_interval = False

        # If the last index could not be defined, set it to the last index of the data
        if last_record_index is None:
            last_record_index = len(self._data.index)

        starting_index = first_record_index
        delta_ts = []
        for t_index in range(first_record_index, last_record_index):
            t0 = datetime.strptime(self._data['Time'].iloc[starting_index], self._time_fmt)
            t1 = datetime.strptime(self._data['Time'].iloc[t_index], self._time_fmt)
            delta_t = (t1 - t0).seconds
            delta_ts.append((t_index, delta_t))

            if delta_t > self._record:
                for r_index, delta_t in delta_ts:
                    if delta_t > self._record:
                        self._record_intervals.append((starting_index, r_index))
                        break
                starting_index = t_index
                delta_ts = []

    @ property
    def parameters(self):
        """Returns the global parameters for the pig.

        This is the first data block stored in the csv file.

        Returns:
            collections.OrderedDict: the pig's parameters.
        """

        return self._parameters

    @property
    def record(self):
        """Return the current record intervals (if any).

        Returns:
            int: the record interval value in minutes.
        """

        if self._record is not None:
            return self._record//60
        else:
            return self._record

    @property
    def record_intervals(self):
        """Return the current record intervals (if any).

        Returns:
            list of 2-tuples: the record inervals.
        """

        return self._record_intervals

    @property
    def record_times(self):
        """Return the starting and ending times of each record interval

        Returns:
            list of 2-tuples: the list of starting and ending time for each record interval
        """

        record_times = [(self._data['Time'].iloc[start], self._data['Time'].iloc[end-1]) for start, end in self._record_intervals]

        return record_times

    @property
    def t_final_interval_index(self):
        """Returns the index of the interval which contains t_final.

        Returns:
            int: the index
        """

        return self.t_interval_index(self._parameters['t_final'])

    def t_interval_index(self, time):
        """Returns the index of the interval which contains a given time.

        Returns:
            int: the index
        """

        record_times = self.record_times

        for index, (_, ending) in enumerate(record_times):
            delta_t = datetime.strptime(time, self._time_fmt) - datetime.strptime(ending, self._time_fmt)
            if delta_t.seconds == 0:
                return index
            if delta_t.days < 0:
                return index

        return len(record_times) - 1

    @property
    def t_initial_interval_index(self):
        """Returns the index of the interval which contains t_initial.

        Returns:
            int: the index
        """

        return self.t_interval_index(self._parameters['t_initial'])

    @property
    def timeline(self):
        """Build the timeline using the record value and the list of intervals.

        Returns:
            list of str: the timeline
        """

        # if no record intervals have been defined yet, returns an empty list
        if not self._record_intervals:
            return []

        initial_time = self.t_initial_interval_index

        record_in_minutes = int(self._record/60)

        timeline = []
        for i in range(-initial_time, len(self._record_intervals)-initial_time):
            time = i*record_in_minutes
            hours = abs(time)//60
            minutes = abs(time) % 60
            minus = '' if time >= 0.0 else '-'
            timeline.append("{}{}h{:02d}".format(minus, hours, minutes))

        return timeline

    def write_summary(self, filename, selected_property='APs'):
        """Write the summay about the statistics for a selected property to an excel file.

        Args:
            filename (str): the excel filename
            selected_property (str): the selected property for which the summary will be written.
        """

        stats = self.get_descriptive_statistics(selected_property)
        if not stats:
            return

        workbook = openpyxl.Workbook()

        # Remove the first empty sheet created by default
        workbook.remove_sheet(workbook.get_sheet_by_name('Sheet'))

        workbook.create_sheet('pig')
        worksheet = workbook.get_sheet_by_name('pig')

        worksheet.cell(row=1, column=11).value = 'Selected property'
        worksheet.cell(row=2, column=11).value = selected_property
        worksheet.cell(row=1, column=1).value = 'Interval'
        worksheet.cell(row=1, column=2).value = 'Average'
        worksheet.cell(row=1, column=3).value = 'Std Dev'
        worksheet.cell(row=1, column=4).value = 'Median'
        worksheet.cell(row=1, column=5).value = '1st quartile'
        worksheet.cell(row=1, column=6).value = '3rd quartile'
        worksheet.cell(row=1, column=7).value = 'Skewness'
        worksheet.cell(row=1, column=8).value = 'kurtosis'

        for i, interval in enumerate(stats['intervals']):

            worksheet.cell(row=i+2, column=1).value = interval
            worksheet.cell(row=i+2, column=2).value = stats['averages'][i]
            worksheet.cell(row=i+2, column=3).value = stats['stddevs'][i]
            worksheet.cell(row=i+2, column=4).value = stats['medians'][i]
            worksheet.cell(row=i+2, column=5).value = stats['1st quantiles'][i]
            worksheet.cell(row=i+2, column=6).value = stats['3rd quantiles'][i]
            worksheet.cell(row=i+2, column=7).value = stats['skewnesses'][i]
            worksheet.cell(row=i+2, column=8).value = stats['kurtosis'][i]

        try:
            workbook.save(filename)
        except PermissionError as error:
            logging.error(str(error))
            return


if __name__ == '__main__':

    reader = PiCCO2FileReader(sys.argv[1])
    reader.set_record_interval(('00:00:00', '6:15:00', 2))
    print(reader.record_times)
    print(reader.t_initial_interval_index)
    print(reader.get_t_final_index())
    print(reader.timeline)
    print(reader.get_descriptive_statistics())
