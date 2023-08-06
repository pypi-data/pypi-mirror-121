# coding: utf-8

# pip install openpyxl==3.0.9
import openpyxl


def read_file(path_file):
    content_list = []
    with open(path_file, encoding='utf-8') as f:
        for i in f:
            row = i.strip()
            content_list.append(row)
    return content_list


def read_excel(file_excel, sheet_name):
    wb = openpyxl.load_workbook(file_excel)
    sheet_names = wb.get_sheet_names()
    if sheet_name not in sheet_names:
        ws = wb.active
    else:
        ws = wb[sheet_name]

    max_row = ws.max_row
    max_col = ws.max_column

    result = []
    for row_index in range(max_row):
        row_index = row_index + 1

        row_list = []
        for col_index in range(max_col):
            col_index = col_index + 1

            cell_value = ws.cell(row=row_index, column=col_index).value
            row_list.append(cell_value)
        result.append(row_list)
    wb.close()
    return result

    @keyword
    def to_excel(self, file_excel, data_list):

        data_list = list(data_list)

        wb = openpyxl.Workbook()
        ws = wb.active

        for row_index, row_list in enumerate(data_list):
            row_index = row_index + 1

            for col_index, cell_value in enumerate(row_list):
                col_index = col_index + 1

                ws.cell(row=row_index, column=col_index, value=cell_value)

        wb.save(file_excel)
